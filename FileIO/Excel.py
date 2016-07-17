#-*- coding: utf-8 -*-

import xlrd
import openpyxl
from openpyxl import Workbook

def cellListToUniList(cellList):
    uniList = []
    for row in cellList:
        rowUni = []
        for ele in row:
            eleUni = ele.value
            rowUni.append(eleUni)
        uniList.append(rowUni)
            
    return uniList


# input dt: string, string
def excelRead(filepath, sheetname, fromRow=1):
    workbook = xlrd.open_workbook(filepath)
    worksheet = workbook.sheet_by_name(sheetname)
    
    num_rows = worksheet.nrows -1
    curr_row = -1
    result = []
    
    while curr_row+fromRow-1 < num_rows:
        curr_row += 1
        row = worksheet.row(curr_row+fromRow-1)
        result.append(row)
    
    resultUni = cellListToUniList(result)
    
    return resultUni
# output dt: list of Cell type



def excelWriteOnExistingFile(filepath, sheetname, columnNum, insert, woFirst=1): 
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name(sheetname)
    
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.active
    
    end_rows = ws.nrows
    curr_row = 0
    insert_rows = len(insert) 
    colInd = len(insert[0])
        
    colNum = columnNum    
    asciiNum = ord(colNum)
    
    indx = 0
    colDigit=0
    
    while indx < colInd:
        curr_row=0
        while curr_row < insert_rows-woFirst:
            curr_row += 1

            try:
                worksheet[colNum+str(end_rows+curr_row)] = insert[curr_row-1+woFirst][indx]
            except(TypeError):
                print('Type Error - '+str(indx))
                
        
        asciiNum += 1
        if asciiNum > 90:
            colDigit+=1
            asciiNum -= 26
        if colDigit == 0:
            colNum=chr(asciiNum)
        else:
            colNum=chr(colDigit+64)+chr(asciiNum)
            
            
        indx+=1     
            
                 
    workbook.save(filepath)
    print('saved successfully in existing file!')
    
def excelWriteNewFile(filepath, sheetname, insertList):
    '''
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheetname)
    
    i=0
    while i<len(insert[0]):
        j=0
        while j<len(insert):
            print(insert[j][i])
            #print(type(insert[j][i]))
            ws.write(j, i, unicode(insert[j][i]))
            j+=1
        i+=1
        
    wb.save(filepath)
    '''
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    
    i1 = 0
    i2 = ord('A')
    while i1<len(insertList[0]):
        j=0
        k=j+1
        while j<len(insertList):
            #print(type(insertList[j][i1]))
            if str(type(insertList[j][i1])) == "<class 'xlrd.sheet.Cell'>":
                ws[chr(i2)+str(k)] = insertList[j][i1].value
            else:
                ws[chr(i2)+str(k)] = insertList[j][i1].encode('utf-8')
            j+=1
            k+=1
        i1+=1
        i2+=1
        
    wb.save(filepath)
    print('saved successfully in a new file!')
    
def xlsToXlsx(filepath, sheetname):
    xlsList = excelRead(filepath, sheetname)
    
    resultList = engToKor(filepath, sheetname)
    engFilepath = resultList[0]+'.xlsx'
    engSheetname = resultList[1]
    
    excelWriteNewFile(engFilepath, engSheetname, xlsList)
    
    fileInfoList = [str(engFilepath), engSheetname]
    return fileInfoList 
    
def engToKor(filepath, sheetname):
    sale = '매매'
    rent = '전월'
    
    # remove last 4 chars (.xls)
    filepath = filepath[:-4]
    
    firstPart = filepath[0:6]
    secondPart = filepath[6:8]
    thirdPart1 = filepath[8:]
    thirdPart2 = filepath[9:]
    
    
    def ifType(thirdPart):
        apartment = '아파트'
        detached = '단독_다가구'
        tenement = '연립_다세대'
        
        if thirdPart == apartment.decode('utf-8'):
            thirdPart = 'Apartment'
        elif thirdPart == detached.decode('utf-8'):
            thirdPart = 'Detached'
        elif thirdPart == tenement.decode('utf-8'):
            thirdPart = 'Tenement'
        return thirdPart
    
    if secondPart == sale.decode('utf-8'):
        secondPart = 'Sale'
        thirdPart = ifType(thirdPart1)
    elif secondPart == rent.decode('utf-8'):
        secondPart = 'Rent'
        thirdPart = ifType(thirdPart2)
    
    resultFilename = firstPart+secondPart+thirdPart
    
    seoul = '서울'
    busan = '부산'
    if sheetname == seoul.decode('utf-8'):
        resultSheetname = 'Seoul'
    elif sheetname == busan.decode('utf-8'):
        resultSheetname = 'Busan'
            
    return [resultFilename, resultSheetname] 

    
    
    